from openpyxl import load_workbook

workbook1 = input('输入处理的目标表格（请不要打开时运行这个程序）')
wb1 = load_workbook(workbook1)
sheets = wb1.sheetnames
ws = wb1[sheets[0]]
# ws1 = wb1['天赋']
target = []
markpos = ws.max_column+1


def pack(value):
    if value != 'None':
        value = '<p style="white-space: pre-wrap;">' + str(value) + '</p>'

    else:
        value = '<p style="white-space: pre-wrap;"></p>'
    return value

def processSheet(ws):
    # print(markpos)
    for i in range(1, ws.max_row + 1):
        flag=0
        for j in range((i + 1), ws.max_row + 1):
            if ws.cell(i, markpos).value != '已合并' and ws.cell(i, 1).value == ws.cell(j, 1).value and ws.cell(i, 1).value != None :
                for p in range(2, markpos):
                    if flag == 0:
                        ws.cell(i, p).value=pack(ws.cell(i, p).value)
                flag =1
                ws.cell(j, markpos).value = '已合并'
                for p in range(2, markpos):
                    ws.cell(i, p).value = str(ws.cell(i, p).value) + pack(str(ws.cell(j, p).value))
    createList(ws)

def createList(ws):
    for i in range(2,ws.max_row+1):
        if ws.cell(i, ws.max_column).value != '已合并' and  ws.cell(i, 1).value != None and  ws.cell(i, 2).value != None:
            temp = []
            for j in range(1,markpos):
                # print(ws.cell(i,j).value)
                temp.append(str(ws.cell(i,j).value))
            target.append(temp)
    # print('target='+str(target))

def writeSheet():
    new_ws = wb1.create_sheet('Sheet1')
    for r in range(0,ws.max_column):
        new_ws.cell(1,r+1).value = ws.cell(1,r+1).value
    for i in range(0,len(target)):
        for j in range(0,len(target[i])):
            # print(target[i][j])
            new_ws.cell(i+2,j+1).value = target[i][j]
# processSheet(ws1)
# processSheet(ws1)

for i in sheets:
    processSheet(wb1[i])
writeSheet()
wb1.save(workbook1)
